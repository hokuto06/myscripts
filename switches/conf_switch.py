from openpyxl import load_workbook, Workbook
from switchesClass import  Switch,ArubaSwitch,CiscoSwitch,BrocadeSwitch


# Script para procesar la lista de switches
def process_switches(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    if "Neighbors" not in wb.sheetnames:
        wb.create_sheet("Neighbors")
    ws_output = wb["Neighbors"]
    if ws_output.max_row == 1:
        ws_output.append(["Switch Hostname", "MAC Address", "Interface", "Device Hostname"])

    for row in ws.iter_rows(min_row=2, values_only=True):
        ip, username, password, brand = row
        print(ip, username,password, brand)
        if brand.lower() == "cisco":
            switch = CiscoSwitch(ip, username, password)
        elif brand.lower() == "aruba":
            switch = ArubaSwitch(ip, username, password)
        elif brand.lower() == "brocade":
            switch = BrocadeSwitch(ip, username, password)
        else:
            continue

        switch.connect()
        neighbors = switch.get_neighbors()
        for neighbor in neighbors:
            ws_output.append(neighbor)
        switch.disconnect()

    wb.save(output_file)


process_switches("switches.xlsx", "aps_on_switches.xlsx")
