import os
import json
from vszClass import connectVsz
from pymongo import MongoClient
import pandas as pd
from pprint import pprint
from openpyxl import load_workbook
from tabulate import tabulate
import argparse
import ipaddress
# from datetime import datetime
client = MongoClient("mongodb://localhost:27017/")  # Asegúrate de que MongoDB esté corriendo
db = client["vsz_db"]  # Nombre de tu base de datos
collection = db["vsz_devices"]  # Colección donde guardarás los dispositivos
db_controller = client["prueba"]
controller_collection = db_controller["mainController_devices"]


def saveOnExcel(rows):
    df = pd.DataFrame(rows)
    output_file = "devices_output.xlsx"
    df.to_excel(output_file, index=False)
    print(f"Archivo guardado como {output_file}")

CONFIG_FILE = "controller_config.json"

def save_controller_config(ip, username, password):
    config = {"ip": ip, "username": username, "password": password}
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f)
    print(f"Configuración guardada en {CONFIG_FILE}")

def load_controller_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    return None

def get_controller_instance():
    config = load_controller_config()

    if not config:
        # Si no hay configuración almacenada, solicita los datos
        ip = input("Ingresa la IP del controlador: ")
        username = input("Ingresa el usuario: ")
        password = input("Ingresa la contraseña: ")
        save_controller_config(ip, username, password)
        config = {"ip": ip, "username": username, "password": password}
    # Crear la instancia del controlador
    return connectVsz(
        ipController=config["ip"],
        userController=config["username"],
        passwordController=config["password"]
    )

# vsz = connectVsz("192.168.188.10")

def set_ap_on_vsz():
    vsz.config_ap('3C:46:A1:2B:14:70', device_description="prueba")

# set_ap_on_vsz()

def save_devices_on_vsz(aps):
    if aps:
        collection.insert_many(aps)

def get_devices_from_vsz():
    # vsz = connectVsz('192.168.188.10')

    aps = vsz.get_all_devices('test')
    
    save_devices_on_vsz(aps)

    pprint(aps)

# get_devices_from_vsz()

#get zones 
def get_get_zones_from_vsz():

    aps = vsz.get_zones()

# get_get_zones_from_vsz()

#Obtiene una lista de todos los devices del vsz y luego hace una consulta completa de cada device
def get_devices_from_vsz_single():
    # vsz = connectVsz('192.168.188.10')

    aps = vsz.get_all_devices('test')
    
    save_devices_on_vsz(aps)

    print(tabulate(aps))
    # pprint(aps)

#
# get_devices_from_vsz_single()

def get_simple_devices():
    aps = vsz.get_all_devices_single()
    pprint(aps)
    print(len(aps))

# get_simple_devices()

def read_from_excel():
    archivo_excel = 'aps_ibis.xlsx'
    #0 name
    #1 ip
    #2 mac
    #5 hab
    #6 cobertura
    #7 nueva ip 
    wb = load_workbook(archivo_excel)
    hoja = wb.active
    # vsz = connectVsz('ruckus1.gocloud1.com')
    for fila in hoja.iter_rows(min_row=1, max_col=hoja.max_column, values_only=True):
        # print("Contenido de la fila:", fila)        
        print(fila[2])
        # vsz.config_full_ap(fila[1],fila[0], fila[7], "255.255.255.0", "10.9.21.1", "Hab "+str(fila[5]))
        vsz.config_full_ap(fila[1],fila[0], "ASD", "255.255.255.0", "10.9.21.1", "Hab ")
        # break 

# read_from_excel()

def read_excel():
    archivo_excel = 'aps_ibis.xlsx'
    wb = load_workbook(archivo_excel)
    hoja = wb.active
    vsz = connectVsz('ruckus1.gocloud1.com')
    for fila in hoja.iter_rows(min_row=1, max_col=hoja.max_column, values_only=True):
        print("Hab "+str(fila[5]))

# read_excel()

# Guarda en un excel todos los datos obteneidos de la coleccion "vsz"
def insert_devices_on_excel():
    cursor = collection.find({})

    list_of_docs = list(cursor)

    df = pd.DataFrame(list_of_docs)
    print(list_of_docs)
    # columns_of_interest = ['name', 'mac','ip_device','serial', 'model', 'description','nodo_padre','port_uplink']
    
    # Asegúrate de que solo incluya esas columnas y en ese orden
    # df = df[columns_of_interest]

    df.to_excel('output.xlsx', index=False, engine='openpyxl')

    print("Datos guardados en 'output.xlsx'")

# insert_devices_on_excel()

def handle_set():
    file = "setupfile.xlsx"

    #Llamar funcion para Crear instancia
    controller = get_controller_instance()

    try:
        # Leer el archivo Excel sin encabezados
        df = pd.read_excel(file, engine="openpyxl", header=None)

        # Verificar que las columnas necesarias existan por índice
        required_indices = [1]  # A=0, E=4, K=10
        if max(required_indices) >= df.shape[1]:
            print(f"Error: El archivo debe contener al menos {max(required_indices) + 1} columnas.")
            return

        # Iterar por cada fila del archivo Excel
        print('''
╔══════════════════════════════════════════════╗
║              Formato del Excel               ║
╠══════════════════════════════════════════════╣
║ Nombre de la planilla: setupfile.xlsx        ║
║                                              ║
║ Columnas:                                    ║
║     A = MAC                                  ║
║     B = IP                                   ║
║     C = NAME                                 ║
║     D = DESCRIPTION                          ║
╚══════════════════════════════════════════════╝
        ''')
        modify_ip = input("¿Configurar IP? (y/n): ").strip().lower() == 'y'
        if modify_ip:
            try:
                device_netmask = input("Ingrese el netmask: ").strip()
                device_gateway = input("Ingrese el gateway: ").strip()
                device_dns = input("Ingrese el DNS: ").strip()

                # Validar que todos los parámetros sean ingresados
                if not device_netmask or not device_gateway or not device_dns:
                    raise ValueError("Todos los parámetros (netmask, gateway, DNS) son obligatorios.")

                # Validar formato de gateway, DNS e IP netmask
                try:
                    ipaddress.IPv4Address(device_gateway)
                    ipaddress.IPv4Address(device_dns)
                    ipaddress.IPv4Network(f"0.0.0.0/{device_netmask}")  # Validar máscara
                except (ipaddress.AddressValueError, ipaddress.NetmaskValueError):
                    raise ValueError("Formato inválido en gateway, DNS o máscara de red.")

            except ValueError as e:
                print(f"Error: {e}")
                print("El script se cancela debido a parámetros inválidos.")
                exit(1)


        modify_name = input("¿Configurar NAME? (y/n): ").strip().lower() == 'y'
        modify_description = input("¿Configurar DESCRIPTION? (y/n): ").strip().lower() == 'y'

        for index, row in df.iterrows():
            device_mac = row[0]  # Columna A (índice 0)
            device_ip = row[1] if not pd.isna(row[1]) else None  # Columna B (índice 1)
            device_name = row[2] if not pd.isna(row[2]) else None  # Columna C (índice 2)
            device_description = row[3] if len(row) > 3 and not pd.isna(row[3]) else None  # Columna D opcional

            if pd.isna(device_mac):
                print(f"Fila {index + 1}: MAC no válida. Saltando esta fila.")
                continue

            print(f"\nConfigurando dispositivo MAC: {device_mac}")
            if modify_ip and device_ip:
                print(f"  -> Configurando IP: {device_ip}")
                print(f"     Netmask: {device_netmask}, Gateway: {device_gateway}, DNS: {device_dns}")
            if modify_name and device_name:
                print(f"  -> Configurando Name: {device_name}")
            if modify_description and device_description:
                print(f"  -> Configurando Description: {device_description}")

            # Configuración del dispositivo
            controller.config_ap(
                device_mac,
                device_name=device_name if modify_name else None,
                device_description=device_description if modify_description else None,
                device_ip=device_ip if modify_ip else None,
                device_netmask=device_netmask if modify_ip else None,
                device_gateway=device_gateway if modify_ip else None,
                device_dns=device_dns if modify_ip else None
            )

    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")


def update_devices_uplinks():
    cursor = controller_collection.find({"deviceType":"switch"})
    aps = collection.find({})
    list_of_aps = list(aps)
    list_of_docs = list(cursor)

    for aps in list_of_aps:
        # print(aps['mac'])        
        
        for switch in list_of_docs:
            nodo_padre = 'none'
            port_uplink = 'none'
            # break
            for ifc, mac in switch["clientes"].items():
                if mac.upper() == aps['mac']:
                    print(mac, aps['name'])
                    # print(aps['_id'])
                    nodo_padre = switch["deviceName"]
                    port_uplink = ifc
                    new_values = { "$set": {'nodo_padre':nodo_padre , 'port_uplink':port_uplink}}
                    filter = {'_id':aps['_id']}
                    collection.update_one(filter, new_values) 
                    break


# update_devices_uplinks()
# insert_devices_on_excel()

def main():
    parser = argparse.ArgumentParser(description="Script para gestionar dispositivos UniFi.")
    parser.add_argument("--get", action="store_true", help="Obtener lista de dispositivos del controlador.")
    parser.add_argument("--set", action="store_true", help="Configurar dispositivos usando un archivo Excel.")
    parser.add_argument("--h", action="store_true", help="Mostrar las opciones disponibles.")
    args = parser.parse_args()

    if args.h:
        parser.print_help()

    elif args.get:
        get_devices_from_vsz_single()

    elif args.set:
        handle_set()

    else:
        print("Opción no válida. Usa --h para ver las opciones.")

if __name__ == "__main__":
    main()