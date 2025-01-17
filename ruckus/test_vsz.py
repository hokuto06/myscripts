import os
import sys
# sys.path.append('/home/hokuto/devicescontroller/mainController')
from vszApi import connectVsz
from pymongo import MongoClient
import pandas as pd
from pprint import pprint
from openpyxl import load_workbook
from tabulate import tabulate
from datetime import datetime
client = MongoClient("mongodb://localhost:27017/")  # Asegúrate de que MongoDB esté corriendo
db = client["vsz_db"]  # Nombre de tu base de datos
collection = db["vsz_devices"]  # Colección donde guardarás los dispositivos
db_controller = client["prueba"]
controller_collection = db_controller["mainController_devices"]
vsz = connectVsz("192.168.188.10")

def set_ap_on_vsz():
    vsz.config_ap('3C:46:A1:2B:14:70', device_description="prueba")

set_ap_on_vsz()

def save_devices_on_vsz(aps):
    if aps:
        collection.insert_many(aps)

def get_devices_from_vsz():
    # vsz = connectVsz('192.168.188.10')

    aps = vsz.get_all_devices('test')
    
    save_devices_on_vsz(aps)

    pprint(aps)

# get_devices_from_vsz()

#get zones 
def get_get_zones_from_vsz():

    aps = vsz.get_zones()

# get_get_zones_from_vsz()

#Obtiene una lista de todos los devices del vsz y luego hace una consulta completa de cada device
def get_devices_from_vsz_single():
    # vsz = connectVsz('192.168.188.10')

    aps = vsz.get_all_devices('test')
    
    save_devices_on_vsz(aps)

    print(tabulate(aps))
    # pprint(aps)

#
# get_devices_from_vsz_single()

def get_simple_devices():
    aps = vsz.get_all_devices_single()
    pprint(aps)
    print(len(aps))

# get_simple_devices()

def read_from_excel():
    archivo_excel = 'aps_ibis.xlsx'
    #0 name
    #1 ip
    #2 mac
    #5 hab
    #6 cobertura
    #7 nueva ip 
    wb = load_workbook(archivo_excel)
    hoja = wb.active
    # vsz = connectVsz('ruckus1.gocloud1.com')
    for fila in hoja.iter_rows(min_row=1, max_col=hoja.max_column, values_only=True):
        # print("Contenido de la fila:", fila)        
        print(fila[2])
        # vsz.config_full_ap(fila[1],fila[0], fila[7], "255.255.255.0", "10.9.21.1", "Hab "+str(fila[5]))
        vsz.config_full_ap(fila[1],fila[0], "ASD", "255.255.255.0", "10.9.21.1", "Hab ")
        # break 

# read_from_excel()

def read_excel():
    archivo_excel = 'aps_ibis.xlsx'
    wb = load_workbook(archivo_excel)
    hoja = wb.active
    vsz = connectVsz('ruckus1.gocloud1.com')
    for fila in hoja.iter_rows(min_row=1, max_col=hoja.max_column, values_only=True):
        print("Hab "+str(fila[5]))

# read_excel()

# Guarda en un excel todos los datos obteneidos de la coleccion "vsz"
def insert_devices_on_excel():
    cursor = collection.find({})

    list_of_docs = list(cursor)

    df = pd.DataFrame(list_of_docs)
    print(list_of_docs)
    # columns_of_interest = ['name', 'mac','ip_device','serial', 'model', 'description','nodo_padre','port_uplink']
    
    # Asegúrate de que solo incluya esas columnas y en ese orden
    # df = df[columns_of_interest]

    df.to_excel('output.xlsx', index=False, engine='openpyxl')

    print("Datos guardados en 'output.xlsx'")

# insert_devices_on_excel()

def update_devices_uplinks():
    cursor = controller_collection.find({"deviceType":"switch"})
    aps = collection.find({})
    list_of_aps = list(aps)
    list_of_docs = list(cursor)

    for aps in list_of_aps:
        # print(aps['mac'])        
        
        for switch in list_of_docs:
            nodo_padre = 'none'
            port_uplink = 'none'
            # break
            for ifc, mac in switch["clientes"].items():
                if mac.upper() == aps['mac']:
                    print(mac, aps['name'])
                    # print(aps['_id'])
                    nodo_padre = switch["deviceName"]
                    port_uplink = ifc
                    new_values = { "$set": {'nodo_padre':nodo_padre , 'port_uplink':port_uplink}}
                    filter = {'_id':aps['_id']}
                    collection.update_one(filter, new_values) 
                    break


# update_devices_uplinks()
# insert_devices_on_excel()

