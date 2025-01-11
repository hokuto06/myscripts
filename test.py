from openpyxl import load_workbook
import os
from pprint import pprint

def _read_excel():
    file_path = 'FS_NHPANORAMA.xlsx'

    if not os.path.exists(file_path):
        print(f"Error: The file '{file_path}' does not exist.")
        return []

    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return []

    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))

    for index, row in enumerate(data):
        if len(row) >= 5:
            try:
                print(f"{row[0]}{row[1]}{row[2]}-{row[3]}-{row[4]}")
                value = f"{row[0]}{row[1]}{row[2]}-{row[3]}-{row[4]}"
                worksheet.cell(row=int(index)+1, column=6, value=value)
            except TypeError as e:
                print(f"Error concatenating row data: {e}")
        else:
            print("Row does not have enough elements")

    try:
        workbook.save(file_path)
        print(f"Datos escritos en la primera columna de la segunda hoja en '{file_path}' exitosamente.")
    except Exception as e:
        print(f"Error al guardar el workbook: {e}")

    workbook.close()
    return data


# import xlwt

# def write_to_excel(file_path, data):
#     workbook = xlwt.Workbook()
#     sheet = workbook.add_sheet('Sheet1')

#     # Escribir datos en el archivo Excel
#     for row_idx, row_data in enumerate(data):
#         for col_idx, col_value in enumerate(row_data):
#             sheet.write(row_idx, col_idx, col_value)

#     # Guardar el archivo Excel
#     workbook.save(file_path)

# if __name__ == "__main__":
#     # Ejemplo de datos a escribir
#     data = [
#         ['Nombre', 'Edad', 'Ciudad'],
#         ['Juan', 30, 'Madrid'],
#         ['María', 25, 'Barcelona'],
#         ['Carlos', 35, 'Valencia'],
#     ]

#     # Ruta del archivo donde se guardará el Excel
#     file_path = 'output.xls'

#     # Llamar a la función para escribir en el Excel
#     write_to_excel(file_path, data)

#     print(f"Archivo Excel guardado en: {file_path}")



data = _read_excel()
